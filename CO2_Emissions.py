import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage

# Load CSV
df = pd.read_csv("co2_emission.csv")

# Clean column names
df.columns = df.columns.str.strip()

# Set exact column name (as seen in your file)
emission_col = "Annual CO₂ emissions (tonnes )"

# Select and clean relevant data
df_cleaned = df[["Entity", "Year", emission_col]].dropna()
df_cleaned["Year"] = pd.to_numeric(df_cleaned["Year"], errors="coerce")
df_cleaned = df_cleaned.dropna(subset=["Year"])

# Get top 10 emitters
top_emitters = (
    df_cleaned.groupby("Entity")[emission_col]
    .sum()
    .sort_values(ascending=False)
    .head(10)
    .index.tolist()
)

# Plot 1: India's CO2 emissions trend
india_data = df_cleaned[df_cleaned["Entity"] == "India"]
plt.figure(figsize=(10, 5))
plt.plot(india_data["Year"], india_data[emission_col], marker="o", color='green')
plt.title("CO₂ Emissions Trend – India")
plt.xlabel("Year")
plt.ylabel("CO₂ Emissions (tonnes)")
plt.grid(True)
plt.tight_layout()
plt.savefig("india_emissions_trend.png")
plt.close()

# Plot 2: Top 10 emitters
total_emissions = (
    df_cleaned[df_cleaned["Entity"].isin(top_emitters)]
    .groupby("Entity")[emission_col]
    .sum()
    .sort_values(ascending=False)
)

plt.figure(figsize=(10, 6))
sns.barplot(x=total_emissions.values, y=total_emissions.index, palette="Reds_r")
plt.title("Top 10 CO₂ Emitting Countries (Total)")
plt.xlabel("Total CO₂ Emissions (tonnes)")
plt.ylabel("Country")
plt.tight_layout()
plt.savefig("top_emitters.png")
plt.close()

# Create Excel Dashboard
wb = Workbook()
ws = wb.active
ws.title = "CO₂ Dashboard"
ws["A1"] = "CO₂ Emissions Dashboard"
ws["A2"] = "1. Line Chart – India"
ws["A20"] = "2. Bar Chart – Top 10 Emitters"

# Insert charts into Excel
img1 = ExcelImage("india_emissions_trend.png")
img2 = ExcelImage("top_emitters.png")
img1.anchor = "A3"
img2.anchor = "A21"
ws.add_image(img1)
ws.add_image(img2)

# Save
wb.save("CO2_Emissions_Dashboard.xlsx")
print("✅ Dashboard created successfully: CO2_Emissions_Dashboard.xlsx")
